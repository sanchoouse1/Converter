from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from copy import deepcopy
import json
import sys
import warnings
import os
#import inspect
import traceback


def convert_to_letter_column(num_column):
    return get_column_letter(num_column)


def write_question_to_arr(
        sheet,
        variant_start,
        variant_end,
        row,
        question_dictionary_copy,
        questions,
        indexes_of_question
        ):
    for key, value_of_key in indexes_of_question.items():
        value_of_key = convert_to_letter_column(value_of_key)
        question_dictionary_copy[key] = sheet[f'{value_of_key}{row}'].value

    for variant in range(variant_start, variant_end + 1):
        variant = convert_to_letter_column(variant)
        if (sheet[f'{variant}{row}'].value != None):
            question_dictionary_copy['variants_of_answer'].append(sheet[f'{variant}{row}'].value)

    questions.append(question_dictionary_copy)


def convert_main():
    # Игнорировать предупреждения типа UserWarning
    warnings.simplefilter("ignore", UserWarning)

    if len(sys.argv) != 3:
        print("Пример использования: python convert.py <имя_excel_файла> <имя_результирующего_файла>")
        sys.exit(1)

    if sys.argv[1].split('.')[-1] != 'xlsx':
        print("Конвертируемый файл должен быть формата .xlsx!")
        sys.exit(1)

    if sys.argv[2].split('.')[-1] != 'json':
        print("Результирующий файл должен быть формата .json!")
        sys.exit(1)

    #calling_script_path1 = os.path.dirname(os.path.abspath(inspect.stack()[1][1]))

    path = traceback.extract_stack()[-2].filename.replace('\\\\', '\\')
    parts = path.split('\\')
    result_path = '\\'.join(parts[:-1])

    excel_filename = os.path.join(result_path, sys.argv[1])
    output_filename = os.path.join(result_path, sys.argv[2])

    workbook = load_workbook(excel_filename)
    sheet = workbook.active

    begin_row_index = 4 # С этой строки идут вопросы
    indexes_of_question = {
        "num": 1,
        "text": 2,
        "type_of_question": 3,
        "correct": 14,
        "timer": 15,
        "scores": 16,
        "link_image": 17
    }

    variant_start = 4
    variant_end = 13

    # max_column_index = sheet.max_column # номер последней заполненной колонки, 17
    max_row_index = sheet.max_row # номер последней строки

    question_dictionary = {
        "num": None,
        "text": "",
        "type_of_question": "",
        "variants_of_answer": [], # отдельный цикл
        "correct": [],
        "timer": None,
        "scores": None,
        "link_image": ""
    }

    questions = []

    for row in range(begin_row_index, max_row_index + 1):
        question_dictionary_copy = deepcopy(question_dictionary)
        write_question_to_arr(
            sheet,
            variant_start,
            variant_end,
            row,
            question_dictionary_copy,
            questions,
            indexes_of_question)

    with open(output_filename, 'w', encoding="utf-8") as json_file:
        json_file.write(json.dumps(questions, indent=4, ensure_ascii=False))


if __name__ == "__main__":
    convert_main()
